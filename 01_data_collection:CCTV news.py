#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
CCTV爬虫
点击CCTV首页右上角搜索按钮，搜索关键词，爬取搜索结果页相关新闻的标题、发布时间、引言、作者
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import json
import csv
from typing import List, Dict
import re
import pandas as pd


class CTVTCrawler:
    """CCTV爬虫类"""
    
    def __init__(self, headless: bool = False, wait_time: int = 10):
        """
        初始化爬虫
        
        Args:
            headless: 是否使用无头模式（不显示浏览器窗口）
            wait_time: 页面加载等待时间（秒）
        """
        self.wait_time = wait_time
        self.driver = None
        self.setup_driver(headless)
    
    def setup_driver(self, headless: bool):
        """设置Chrome浏览器驱动"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.implicitly_wait(self.wait_time)
        except Exception as e:
            print(f"初始化浏览器驱动失败: {e}")
            print("请确保已安装Chrome浏览器和chromedriver")
            raise
    
    def click_search_button(self) -> bool:
        """
        访问CCTV首页并点击右上角的搜索按钮
        
        Returns:
            是否成功点击搜索按钮
        """
        try:
            print("正在访问CCTV首页...")
            self.driver.get("https://www.cctv.com")
            time.sleep(3)  # 等待页面加载
            
            print("正在查找搜索按钮...")
            # 尝试多种方式查找搜索按钮
            search_button_selectors = [
                # 通过class查找
                (By.CSS_SELECTOR, ".search-btn, .search-button, .search-icon, [class*='search']"),
                # 通过id查找
                (By.ID, "search-btn, search-button, search-icon"),
                # 通过XPath查找（放大镜图标通常在右上角）
                (By.XPATH, "//a[contains(@class, 'search')]"),
                (By.XPATH, "//button[contains(@class, 'search')]"),
                (By.XPATH, "//i[contains(@class, 'search')]"),
                (By.XPATH, "//span[contains(@class, 'search')]"),
                # 通过文本查找
                (By.XPATH, "//a[contains(text(), '搜索')]"),
                # 通过title属性查找
                (By.XPATH, "//a[@title='搜索']"),
                # 查找右上角的搜索相关元素
                (By.XPATH, "//div[contains(@class, 'header')]//a[contains(@class, 'search')]"),
                (By.XPATH, "//div[contains(@class, 'nav')]//a[contains(@class, 'search')]"),
            ]
            
            search_button = None
            for by, selector in search_button_selectors:
                try:
                    elements = self.driver.find_elements(by, selector)
                    for element in elements:
                        if element.is_displayed():
                            search_button = element
                            print(f"找到搜索按钮（使用选择器: {by}, {selector}）")
                            break
                    if search_button:
                        break
                except:
                    continue
            
            # 如果还没找到，尝试通过JavaScript查找
            if not search_button:
                print("尝试通过JavaScript查找搜索按钮...")
                search_button = self.driver.execute_script("""
                    var elements = document.querySelectorAll('a, button, i, span, div');
                    for(var i = 0; i < elements.length; i++) {
                        var elem = elements[i];
                        var className = elem.className || '';
                        var id = elem.id || '';
                        var text = elem.textContent || elem.innerText || '';
                        var title = elem.title || '';
                        
                        if((className.toLowerCase().includes('search') || 
                            id.toLowerCase().includes('search') ||
                            text.includes('搜索') ||
                            title.includes('搜索')) &&
                           elem.offsetParent !== null) {
                            return elem;
                        }
                    }
                    return null;
                """)
            
            if not search_button:
                print("未找到搜索按钮，尝试直接查找搜索框...")
                # 如果找不到搜索按钮，可能搜索框已经可见，直接查找搜索框
                return self.find_search_input()
            
            # 点击搜索按钮
            try:
                # 滚动到元素可见
                self.driver.execute_script("arguments[0].scrollIntoView(true);", search_button)
                time.sleep(0.5)
                
                # 尝试普通点击
                search_button.click()
            except:
                # 如果普通点击失败，尝试JavaScript点击
                self.driver.execute_script("arguments[0].click();", search_button)
            
            time.sleep(2)  # 等待搜索框展开
            print("成功点击搜索按钮")
            return True
            
        except Exception as e:
            print(f"点击搜索按钮时出错: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def find_search_input(self) -> bool:
        """
        查找搜索输入框
        
        Returns:
            是否找到搜索输入框
        """
        try:
            # 尝试多种方式查找搜索输入框
            search_input_selectors = [
                (By.CSS_SELECTOR, "input[type='search'], input[type='text'][placeholder*='搜索'], input[name*='search'], input[id*='search']"),
                (By.XPATH, "//input[@type='search']"),
                (By.XPATH, "//input[contains(@placeholder, '搜索')]"),
                (By.XPATH, "//input[contains(@class, 'search')]"),
                (By.XPATH, "//input[contains(@id, 'search')]"),
                (By.XPATH, "//input[contains(@name, 'search')]"),
            ]
            
            for by, selector in search_input_selectors:
                try:
                    search_input = self.driver.find_element(by, selector)
                    if search_input.is_displayed():
                        print(f"找到搜索输入框（使用选择器: {by}, {selector}）")
                        return True
                except:
                    continue
            
            return False
        except Exception as e:
            print(f"查找搜索输入框时出错: {e}")
            return False
    
    def search_keyword(self, keyword: str) -> bool:
        """
        在搜索框中输入关键词并搜索
        
        Args:
            keyword: 搜索关键词
            
        Returns:
            是否成功搜索
        """
        try:
            # 先点击搜索按钮（如果还没点击）
            if not self.find_search_input():
                if not self.click_search_button():
                    print("无法打开搜索框，尝试直接访问搜索URL...")
                    return self.search_direct_url(keyword)
            
            # 查找搜索输入框
            search_input_selectors = [
                (By.CSS_SELECTOR, "input[type='search'], input[type='text'][placeholder*='搜索'], input[name*='search'], input[id*='search']"),
                (By.XPATH, "//input[@type='search']"),
                (By.XPATH, "//input[contains(@placeholder, '搜索')]"),
                (By.XPATH, "//input[contains(@class, 'search')]"),
                (By.XPATH, "//input[contains(@id, 'search')]"),
                (By.XPATH, "//input[contains(@name, 'search')]"),
            ]
            
            search_input = None
            for by, selector in search_input_selectors:
                try:
                    elements = self.driver.find_elements(by, selector)
                    for element in elements:
                        if element.is_displayed():
                            search_input = element
                            break
                    if search_input:
                        break
                except:
                    continue
            
            if not search_input:
                print("未找到搜索输入框，尝试直接访问搜索URL...")
                return self.search_direct_url(keyword)
            
            # 清空输入框并输入关键词
            search_input.clear()
            search_input.send_keys(keyword)
            time.sleep(1)
            
            # 查找并点击搜索按钮或按回车
            try:
                # 尝试查找搜索提交按钮
                submit_selectors = [
                    (By.XPATH, "//button[@type='submit']"),
                    (By.XPATH, "//input[@type='submit']"),
                    (By.XPATH, "//button[contains(text(), '搜索')]"),
                    (By.XPATH, "//a[contains(text(), '搜索')]"),
                    (By.CSS_SELECTOR, ".search-submit, .submit-search"),
                ]
                
                submit_button = None
                for by, selector in submit_selectors:
                    try:
                        button = self.driver.find_element(by, selector)
                        if button.is_displayed():
                            submit_button = button
                            break
                    except:
                        continue
                
                if submit_button:
                    submit_button.click()
                else:
                    # 如果找不到提交按钮，按回车键
                    search_input.send_keys(Keys.RETURN)
            except:
                # 如果点击失败，按回车键
                search_input.send_keys(Keys.RETURN)
            
            time.sleep(3)  # 等待搜索结果加载
            print(f"成功搜索关键词: {keyword}")
            return True
            
        except Exception as e:
            print(f"搜索关键词时出错: {e}")
            import traceback
            traceback.print_exc()
            # 如果出错，尝试直接访问搜索URL
            return self.search_direct_url(keyword)
    
    def search_direct_url(self, keyword: str) -> bool:
        """
        直接访问搜索URL（备用方案）
        
        Args:
            keyword: 搜索关键词
            
        Returns:
            是否成功访问
        """
        try:
            from urllib.parse import quote
            encoded_keyword = quote(keyword, safe='')
            search_url = f"https://search.cctv.com/search.php?qtext={encoded_keyword}&type=web"
            print(f"直接访问搜索URL: {search_url}")
            self.driver.get(search_url)
            time.sleep(3)
            return True
        except Exception as e:
            print(f"直接访问搜索URL时出错: {e}")
            return False
    
    def parse_search_results(self, keyword: str = '') -> List[Dict]:
        """
        解析搜索结果页面的新闻列表
        
        Returns:
            新闻信息列表，每个元素包含标题、发布时间、引言、作者
        """
        news_list = []
        
        try:
            # 滚动页面确保所有内容加载
            self.scroll_to_bottom()
            time.sleep(2)
            
            # 获取页面源码
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # 尝试多种选择器来查找新闻条目
            news_selectors = [
                ('div', {'class': re.compile(r'result|item|news|article', re.I)}),
                ('li', {'class': re.compile(r'result|item|news|article', re.I)}),
                ('div', {'class': re.compile(r'search.*result', re.I)}),
                ('a', {'class': re.compile(r'news|article', re.I)}),
            ]
            
            news_items = []
            for tag, attrs in news_selectors:
                items = soup.find_all(tag, attrs)
                if items:
                    print(f"找到 {len(items)} 条新闻（使用选择器: {tag} with {attrs}）")
                    news_items = items
                    break
            
            # 如果还是找不到，尝试更通用的方法
            if not news_items:
                all_divs = soup.find_all(['div', 'li'], class_=True)
                for div in all_divs:
                    if div.find('a') and (div.find('h3') or div.find('h2') or div.find('h1')):
                        news_items.append(div)
                print(f"使用通用方法找到 {len(news_items)} 条可能的新闻")
            
            # 解析每条新闻
            for item in news_items:
                try:
                    news_info = self._parse_news_item(item, keyword=keyword)
                    if news_info.get('title'):  # 至少需要标题
                        news_list.append(news_info)
                except Exception as e:
                    print(f"解析单条新闻时出错: {e}")
                    continue
            
            # 如果还是没找到，尝试从页面文本中提取
            if not news_list:
                print("尝试从页面文本中提取新闻信息...")
                title_links = soup.find_all('a', href=re.compile(r'/news/|/tv/|/video/'))
                for link in title_links[:20]:
                    try:
                        title = link.get_text(strip=True)
                        if title and len(title) > 5:
                            raw_url = link.get('href', '')
                            news_info = {
                                'keyword': keyword,
                                'title': title,
                                'publish_time': self._extract_time_from_element(link.parent),
                                'summary': self._extract_summary_from_element(link.parent),
                                'author': self._extract_author_from_element(link.parent),
                                'url': self._normalize_url(raw_url) if raw_url else ''
                            }
                            if news_info['title']:
                                news_list.append(news_info)
                    except Exception as e:
                        continue
            
            print(f"共解析到 {len(news_list)} 条新闻")
            
        except Exception as e:
            print(f"解析搜索结果时出错: {e}")
            import traceback
            traceback.print_exc()
        
        return news_list
    
    def crawl_all_pages(self, keyword: str = '') -> List[Dict]:
        """
        爬取所有页面的新闻（自动下滑、自动翻页，直到没有下一页按钮）
        
        Args:
            keyword: 当前搜索的关键词
            
        Returns:
            所有页面的新闻信息列表
        """
        all_news = []
        page_num = 1
        
        print(f"\n开始爬取所有页面（自动爬取直到没有下一页）...")
        print("="*80)
        
        while True:
            print(f"\n【第 {page_num} 页】")
            print("-"*80)
            
            # 如果是第一页，先点击"发布时间" -> "一年以内"筛选
            if page_num == 1:
                self.click_publish_time_filter()
            
            # 自动下滑到底部
            print("正在下滑到页面底部...")
            self.scroll_to_bottom()
            time.sleep(1)
            
            # 解析当前页面的新闻
            page_news = self.parse_search_results(keyword=keyword)
            
            if not page_news:
                print(f"第 {page_num} 页未找到新闻，停止爬取")
                break
            
            # 添加到总结果中
            all_news.extend(page_news)
            print(f"第 {page_num} 页获得 {len(page_news)} 条新闻，累计 {len(all_news)} 条")
            
            # 尝试点击下一页
            if not self.click_next_page():
                print(f"无法继续翻页，已到达最后一页")
                break
            
            page_num += 1
            time.sleep(2)  # 页面切换后等待
        
        print("\n" + "="*80)
        print(f"爬取完成！共爬取 {page_num} 页，获得 {len(all_news)} 条新闻")
        print("="*80)
        
        return all_news
    
    def _parse_news_item(self, item, keyword: str = '') -> Dict:
        """解析单条新闻项"""
        news_info = {
            'keyword': keyword,
            'title': '',
            'publish_time': '',
            'summary': '',
            'author': '',
            'url': ''
        }
        
        # 提取标题和链接
        title_elem = item.find(['h1', 'h2', 'h3', 'h4', 'h5'])
        if title_elem:
            news_info['title'] = title_elem.get_text(strip=True)
            # 尝试从标题元素附近查找链接
            link_elem = title_elem.find('a', href=True)
            if not link_elem:
                link_elem = title_elem.find_parent('a', href=True)
            if not link_elem:
                link_elem = item.find('a', href=True)
        
        if not news_info['title']:
            link_elem = item.find('a', href=True)
            if link_elem:
                title = link_elem.get_text(strip=True)
                if len(title) > 100:
                    title_attr = link_elem.get('title', '')
                    if title_attr:
                        title = title_attr.strip()
                news_info['title'] = title
        
        if not news_info['title']:
            title_class_elem = item.find(class_=re.compile(r'title', re.I))
            if title_class_elem:
                news_info['title'] = title_class_elem.get_text(strip=True)
                # 尝试从标题类元素中查找链接
                if not link_elem:
                    link_elem = title_class_elem.find('a', href=True)
        
        # 提取链接URL
        if link_elem:
            raw_url = link_elem.get('href', '')
            if raw_url:
                news_info['url'] = self._normalize_url(raw_url)
        
        # 如果还没有找到链接，尝试其他方式
        if not news_info['url']:
            # 尝试从data-url等属性获取
            for attr in ['data-url', 'data-href', 'data-link']:
                url_attr = item.get(attr, '')
                if url_attr:
                    news_info['url'] = self._normalize_url(url_attr)
                    break
        
        # 提取发布时间
        news_info['publish_time'] = self._extract_time_from_element(item)
        
        # 提取引言（简介）
        news_info['summary'] = self._extract_summary_from_element(item)
        
        # 提取作者
        news_info['author'] = self._extract_author_from_element(item)
        
        return news_info
    
    def _normalize_url(self, url: str) -> str:
        """
        规范化URL，补全相对路径
        
        Args:
            url: 原始URL
            
        Returns:
            规范化后的完整URL
        """
        if not url:
            return ''
        
        url = url.strip()
        
        # 如果已经是完整URL，直接返回
        if url.startswith('http://') or url.startswith('https://'):
            return url
        
        # 处理协议相对URL（以//开头）
        if url.startswith('//'):
            return 'https:' + url
        
        # 处理绝对路径（以/开头）
        if url.startswith('/'):
            return 'https://www.cctv.com' + url
        
        # 其他情况，尝试补全
        if not url.startswith('http'):
            return 'https://www.cctv.com/' + url.lstrip('/')
        
        return url
    
    def _extract_time_from_element(self, element) -> str:
        """从元素中提取发布时间"""
        time_patterns = [
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',
            r'\d{4}年\d{1,2}月\d{1,2}日',
            r'\d{1,2}[-/]\d{1,2}[-/]\d{4}',
            r'\d{4}\.\d{1,2}\.\d{1,2}',
        ]
        
        # 查找时间相关的元素
        time_elem = element.find(['span', 'div', 'p', 'time', 'em', 'i'], 
                                 class_=re.compile(r'time|date|publish|pub', re.I))
        if time_elem:
            text = time_elem.get_text()
            for pattern in time_patterns:
                match = re.search(pattern, text)
                if match:
                    return match.group()
        
        # 查找datetime属性
        time_attr_elem = element.find(attrs={'datetime': True})
        if time_attr_elem:
            datetime_attr = time_attr_elem.get('datetime', '')
            for pattern in time_patterns:
                match = re.search(pattern, datetime_attr)
                if match:
                    return match.group()
        
        # 在整个元素文本中搜索
        text = element.get_text() if hasattr(element, 'get_text') else str(element)
        for pattern in time_patterns:
            match = re.search(pattern, text)
            if match:
                return match.group()
        
        return ''
    
    def _extract_summary_from_element(self, element) -> str:
        """从元素中提取引言（简介）"""
        # 查找简介相关的元素
        summary_elem = element.find(['p', 'div', 'span'], 
                                   class_=re.compile(r'summary|desc|intro|abstract|content', re.I))
        if summary_elem:
            summary = summary_elem.get_text(strip=True)
            if summary and len(summary) > 10:
                return summary[:200]
        
        # 尝试获取所有p标签
        p_elems = element.find_all('p')
        for p_elem in p_elems:
            summary = p_elem.get_text(strip=True)
            if 20 <= len(summary) <= 500:
                return summary[:200]
        
        # 尝试从div中提取
        div_elems = element.find_all('div')
        for div_elem in div_elems:
            if div_elem.find(['h1', 'h2', 'h3', 'h4', 'a']):
                continue
            summary = div_elem.get_text(strip=True)
            if 20 <= len(summary) <= 500:
                return summary[:200]
        
        return ''
    
    def _extract_author_from_element(self, element) -> str:
        """从元素中提取作者"""
        # 查找作者相关的元素
        author_elem = element.find(['span', 'div', 'a', 'em', 'i'], 
                                  class_=re.compile(r'author|writer|reporter|记者|作者', re.I))
        if author_elem:
            author_text = author_elem.get_text(strip=True)
            if author_text and len(author_text) < 50:
                return author_text
        
        # 在整个元素文本中搜索作者关键词
        text = element.get_text() if hasattr(element, 'get_text') else str(element)
        # 查找"记者"、"作者"等关键词后的内容
        author_patterns = [
            r'记者[：:]\s*([^\s，,。]+)',
            r'作者[：:]\s*([^\s，,。]+)',
            r'编辑[：:]\s*([^\s，,。]+)',
            r'来源[：:]\s*([^\s，,。]+)',
        ]
        
        for pattern in author_patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)
        
        return ''
    
    def click_publish_time_filter(self) -> bool:
        """
        在搜索结果页点击"发布时间"筛选，然后选择"一年以内"
        
        Returns:
            是否成功点击筛选
        """
        try:
            print("正在查找并点击'发布时间'筛选...")
            time.sleep(2)  # 等待搜索结果页加载
            
            # 第一步：查找并点击"发布时间"按钮
            publish_time_selectors = [
                # 通过文本查找
                (By.XPATH, "//a[contains(text(), '发布时间')]"),
                (By.XPATH, "//span[contains(text(), '发布时间')]"),
                (By.XPATH, "//button[contains(text(), '发布时间')]"),
                (By.XPATH, "//div[contains(text(), '发布时间')]"),
                (By.XPATH, "//li[contains(text(), '发布时间')]"),
                # 通过链接文本查找
                (By.LINK_TEXT, "发布时间"),
                (By.PARTIAL_LINK_TEXT, "发布时间"),
                # 通过class查找（可能包含filter、sort、time等关键词）
                (By.CSS_SELECTOR, "[class*='filter'], [class*='sort'], [class*='time']"),
            ]
            
            publish_time_button = None
            for by, selector in publish_time_selectors:
                try:
                    elements = self.driver.find_elements(by, selector)
                    for element in elements:
                        text = element.text or element.get_attribute('textContent') or ''
                        if '发布时间' in text:
                            if element.is_displayed():
                                publish_time_button = element
                                print(f"找到'发布时间'按钮（使用选择器: {by}, {selector}）")
                                break
                    if publish_time_button:
                        break
                except:
                    continue
            
            # 如果还没找到，尝试通过JavaScript查找
            if not publish_time_button:
                print("尝试通过JavaScript查找'发布时间'按钮...")
                publish_time_button = self.driver.execute_script("""
                    var elements = document.querySelectorAll('a, button, span, div, li');
                    for(var i = 0; i < elements.length; i++) {
                        var elem = elements[i];
                        var text = elem.textContent || elem.innerText || '';
                        if(text.includes('发布时间') && elem.offsetParent !== null) {
                            return elem;
                        }
                    }
                    return null;
                """)
            
            if not publish_time_button:
                print("未找到'发布时间'按钮，跳过筛选")
                return False
            
            # 点击"发布时间"按钮
            try:
                # 滚动到元素可见
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", publish_time_button)
                time.sleep(0.5)
                
                # 尝试普通点击
                publish_time_button.click()
            except:
                # 如果普通点击失败，尝试JavaScript点击
                self.driver.execute_script("arguments[0].click();", publish_time_button)
            
            # 等待下拉菜单展开，使用WebDriverWait
            print("等待下拉菜单展开...")
            try:
                # 使用WebDriverWait等待"一年以内"选项出现
                wait = WebDriverWait(self.driver, 5)
                # 先等待任何包含"一年"的元素出现
                wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), '一年')]")))
                print("检测到包含'一年'的元素出现")
            except:
                print("WebDriverWait超时，继续尝试查找...")
            
            # 额外等待确保菜单完全展开
            time.sleep(2)
            
            # 打印所有可见的下拉选项用于调试
            try:
                print("正在查找所有可见的下拉选项...")
                all_options = self.driver.find_elements(By.XPATH, "//*[contains(text(), '一年') or contains(text(), '一月') or contains(text(), '一周') or contains(text(), '一天') or contains(text(), '不限')]")
                visible_options = []
                for opt in all_options:
                    if opt.is_displayed():
                        text = opt.text or opt.get_attribute('textContent') or ''
                        if text.strip():
                            visible_options.append(text.strip())
                if visible_options:
                    print(f"找到以下可见选项: {visible_options}")
            except Exception as e:
                print(f"查找选项时出错: {e}")
            
            # 第二步：查找并点击"一年以内"选项
            print("正在查找并点击'一年以内'选项...")
            
            # 首先尝试查找下拉菜单容器，然后在其中查找
            one_year_option = None
            
            # 策略1：在下拉菜单容器中查找
            dropdown_selectors = [
                "ul.dropdown-menu",
                "div.dropdown-menu",
                "ul[class*='dropdown']",
                "div[class*='dropdown']",
                "ul[class*='menu']",
                "div[class*='menu']",
                "ul[class*='list']",
                "div[class*='list']",
                "ul",
                "div[role='menu']",
                "ul[role='menu']"
            ]
            
            for dropdown_selector in dropdown_selectors:
                try:
                    dropdowns = self.driver.find_elements(By.CSS_SELECTOR, dropdown_selector)
                    for dropdown in dropdowns:
                        if dropdown.is_displayed():
                            # 在下拉菜单中查找"一年以内"
                            options = dropdown.find_elements(By.TAG_NAME, "a")
                            options.extend(dropdown.find_elements(By.TAG_NAME, "li"))
                            options.extend(dropdown.find_elements(By.TAG_NAME, "span"))
                            options.extend(dropdown.find_elements(By.TAG_NAME, "div"))
                            
                            for option in options:
                                text = option.text or option.get_attribute('textContent') or ''
                                if '一年以内' in text:
                                    if option.is_displayed():
                                        one_year_option = option
                                        print(f"在下拉菜单中找到'一年以内'选项")
                                        break
                            if one_year_option:
                                break
                    if one_year_option:
                        break
                except:
                    continue
            
            # 策略2：使用XPath直接查找所有包含"一年以内"的元素
            if not one_year_option:
                one_year_selectors = [
                    # 通过文本查找（更精确）
                    (By.XPATH, "//a[normalize-space(text())='一年以内']"),
                    (By.XPATH, "//li[normalize-space(text())='一年以内']"),
                    (By.XPATH, "//span[normalize-space(text())='一年以内']"),
                    (By.XPATH, "//div[normalize-space(text())='一年以内']"),
                    # 包含文本
                    (By.XPATH, "//a[contains(text(), '一年以内')]"),
                    (By.XPATH, "//li[contains(text(), '一年以内')]"),
                    (By.XPATH, "//span[contains(text(), '一年以内')]"),
                    (By.XPATH, "//div[contains(text(), '一年以内')]"),
                    (By.XPATH, "//*[contains(text(), '一年以内')]"),
                    # 通过链接文本查找
                    (By.LINK_TEXT, "一年以内"),
                    (By.PARTIAL_LINK_TEXT, "一年以内"),
                ]
                
                for by, selector in one_year_selectors:
                    try:
                        elements = self.driver.find_elements(by, selector)
                        for element in elements:
                            text = element.text or element.get_attribute('textContent') or ''
                            if '一年以内' in text:
                                # 检查元素是否可见且可点击
                                if element.is_displayed():
                                    # 检查元素是否在视口中
                                    try:
                                        location = element.location
                                        size = element.size
                                        if location['x'] >= 0 and location['y'] >= 0 and size['width'] > 0 and size['height'] > 0:
                                            one_year_option = element
                                            print(f"找到'一年以内'选项（使用选择器: {by}, {selector}）")
                                            break
                                    except:
                                        one_year_option = element
                                        print(f"找到'一年以内'选项（使用选择器: {by}, {selector}）")
                                        break
                        if one_year_option:
                            break
                    except:
                        continue
            
            # 策略3：通过JavaScript查找（更全面）
            if not one_year_option:
                print("尝试通过JavaScript查找'一年以内'选项...")
                one_year_option = self.driver.execute_script("""
                    // 首先查找所有可能的下拉菜单
                    var dropdowns = document.querySelectorAll('ul, div[class*="menu"], div[class*="dropdown"], div[class*="list"]');
                    var found = null;
                    
                    // 在下拉菜单中查找
                    for(var d = 0; d < dropdowns.length; d++) {
                        var dropdown = dropdowns[d];
                        var style = window.getComputedStyle(dropdown);
                        if(style.display !== 'none' && style.visibility !== 'hidden') {
                            var elements = dropdown.querySelectorAll('a, li, span, div');
                            for(var i = 0; i < elements.length; i++) {
                                var elem = elements[i];
                                var text = elem.textContent || elem.innerText || '';
                                if(text.includes('一年以内') || text.includes('一年')) {
                                    var elemStyle = window.getComputedStyle(elem);
                                    if(elemStyle.display !== 'none' && elemStyle.visibility !== 'hidden' && elem.offsetParent !== null) {
                                        found = elem;
                                        break;
                                    }
                                }
                            }
                            if(found) break;
                        }
                    }
                    
                    // 如果还没找到，在整个页面中查找
                    if(!found) {
                        var allElements = document.querySelectorAll('a, button, span, div, li');
                        for(var i = 0; i < allElements.length; i++) {
                            var elem = allElements[i];
                            var text = elem.textContent || elem.innerText || '';
                            if((text.includes('一年以内') || text.includes('一年')) && elem.offsetParent !== null) {
                                var style = window.getComputedStyle(elem);
                                if(style.display !== 'none' && style.visibility !== 'hidden') {
                                    found = elem;
                                    break;
                                }
                            }
                        }
                    }
                    
                    return found;
                """)
            
            if not one_year_option:
                print("未找到'一年以内'选项，跳过筛选")
                # 尝试截图保存当前页面状态用于调试
                try:
                    self.driver.save_screenshot('debug_no_one_year.png')
                    print("已保存页面截图到 debug_no_one_year.png 用于调试")
                except:
                    pass
                return False
            
            # 点击"一年以内"选项
            print("准备点击'一年以内'选项...")
            click_success = False
            
            # 尝试多种点击方式
            click_attempts = [
                ("普通点击", lambda elem: elem.click()),
                ("JavaScript点击", lambda elem: self.driver.execute_script("arguments[0].click();", elem)),
                ("事件触发点击", lambda elem: self.driver.execute_script("""
                    var elem = arguments[0];
                    var event = new MouseEvent('click', {
                        view: window,
                        bubbles: true,
                        cancelable: true
                    });
                    elem.dispatchEvent(event);
                """, elem)),
                ("mousedown+mouseup+click", lambda elem: self.driver.execute_script("""
                    var elem = arguments[0];
                    var events = ['mousedown', 'mouseup', 'click'];
                    events.forEach(function(eventType) {
                        var event = new MouseEvent(eventType, {
                            view: window,
                            bubbles: true,
                            cancelable: true
                        });
                        elem.dispatchEvent(event);
                    });
                """, elem)),
            ]
            
            for method_name, click_method in click_attempts:
                try:
                    print(f"尝试{method_name}...")
                    # 滚动到元素可见
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", one_year_option)
                    time.sleep(0.5)
                    
                    # 先尝试鼠标悬停
                    try:
                        from selenium.webdriver.common.action_chains import ActionChains
                        ActionChains(self.driver).move_to_element(one_year_option).perform()
                        time.sleep(0.3)
                    except:
                        pass
                    
                    # 执行点击
                    click_method(one_year_option)
                    time.sleep(1)  # 等待点击生效
                    
                    # 验证是否点击成功（检查下拉菜单是否关闭或URL是否变化）
                    current_url = self.driver.current_url
                    if 'datepid=5' in current_url or 'datepid=1' in current_url:
                        print(f"{method_name}成功！")
                        click_success = True
                        break
                    else:
                        # 检查下拉菜单是否关闭
                        try:
                            # 如果下拉菜单关闭了，说明点击可能成功了
                            dropdowns = self.driver.find_elements(By.CSS_SELECTOR, 
                                "ul.dropdown-menu, div.dropdown-menu, ul[class*='dropdown']")
                            menu_closed = True
                            for dd in dropdowns:
                                if dd.is_displayed():
                                    menu_closed = False
                                    break
                            if menu_closed:
                                print(f"{method_name}可能成功（下拉菜单已关闭）")
                                click_success = True
                                break
                        except:
                            pass
                            
                except Exception as e:
                    print(f"{method_name}失败: {e}")
                    continue
            
            if not click_success:
                print("所有点击方式都失败，尝试强制点击...")
                # 最后尝试：直接通过JavaScript查找并点击
                try:
                    result = self.driver.execute_script("""
                        var elements = document.querySelectorAll('a, li, span, div');
                        for(var i = 0; i < elements.length; i++) {
                            var elem = elements[i];
                            var text = elem.textContent || elem.innerText || '';
                            if(text.includes('一年以内')) {
                                var style = window.getComputedStyle(elem);
                                if(style.display !== 'none' && style.visibility !== 'hidden' && elem.offsetParent !== null) {
                                    elem.scrollIntoView({behavior: 'smooth', block: 'center'});
                                    setTimeout(function() {
                                        elem.click();
                                    }, 100);
                                    return true;
                                }
                            }
                        }
                        return false;
                    """)
                    if result:
                        print("通过JavaScript强制点击成功")
                        click_success = True
                    else:
                        print("JavaScript强制点击也失败")
                except Exception as e:
                    print(f"JavaScript强制点击出错: {e}")
            
            if not click_success:
                print("警告：无法确认'一年以内'选项是否被点击，继续执行...")
                return False
            
            time.sleep(3)  # 等待筛选结果加载
            
            # 验证是否成功（检查URL）
            print("验证筛选是否成功...")
            time.sleep(1)
            final_url = self.driver.current_url
            if 'datepid=5' in final_url:
                print("筛选成功！URL中包含datepid=5参数（一年以内）")
            elif 'datepid=1' in final_url:
                print("筛选可能成功！URL中包含datepid=1参数")
            else:
                print(f"警告：URL中未找到筛选参数，当前URL: {final_url}")
            
            print("成功点击'发布时间' -> '一年以内'筛选")
            return True
                
        except Exception as e:
            print(f"点击筛选时出错: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def filter_by_year(self, news_list: List[Dict], year: int = 2025) -> List[Dict]:
        """
        筛选指定年份的新闻
        
        Args:
            news_list: 新闻列表
            year: 目标年份，默认为2025
            
        Returns:
            筛选后的新闻列表
        """
        filtered_news = []
        
        for news in news_list:
            publish_time = news.get('publish_time', '').strip()
            
            if not publish_time:
                # 如果发布时间为空，跳过
                continue
            
            # 尝试从时间字符串中提取年份
            year_match = re.search(r'(\d{4})', publish_time)
            if year_match:
                news_year = int(year_match.group(1))
                if news_year == year:
                    filtered_news.append(news)
        
        return filtered_news
    
    def scroll_to_bottom(self, scroll_pause_time: float = 1.0) -> bool:
        """
        滚动页面到底部，检测是否已到底部
        
        Args:
            scroll_pause_time: 每次滚动后的等待时间（秒）
            
        Returns:
            是否已滚动到底部
        """
        try:
            # 获取当前滚动位置和页面高度
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            current_position = self.driver.execute_script("return window.pageYOffset + window.innerHeight")
            
            scroll_count = 0
            max_scrolls = 20  # 增加最大滚动次数
            
            while scroll_count < max_scrolls:
                # 滚动到底部
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(scroll_pause_time)
                
                # 获取新的页面高度和当前位置
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                new_position = self.driver.execute_script("return window.pageYOffset + window.innerHeight")
                
                # 检查是否已到底部（允许5像素的误差）
                if abs(new_position - new_height) < 5:
                    print("已滚动到页面底部")
                    return True
                
                # 如果页面高度没有变化，可能已经到底
                if new_height == last_height:
                    # 再尝试滚动一次确认
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(scroll_pause_time)
                    final_position = self.driver.execute_script("return window.pageYOffset + window.innerHeight")
                    final_height = self.driver.execute_script("return document.body.scrollHeight")
                    if abs(final_position - final_height) < 5:
                        print("已滚动到页面底部")
                        return True
                    break
                
                last_height = new_height
                scroll_count += 1
            
            print(f"滚动完成，共滚动 {scroll_count} 次")
            return True
            
        except Exception as e:
            print(f"滚动页面时出错: {e}")
            return False
    
    def find_next_page_button(self):
        """
        查找下一页按钮
        
        Returns:
            下一页按钮元素，如果找不到则返回None
        """
        try:
            # 尝试多种选择器来查找下一页按钮
            next_page_selectors = [
                # 通过文本查找
                (By.XPATH, "//a[contains(text(), '下一页')]"),
                (By.XPATH, "//a[contains(text(), '下页')]"),
                (By.XPATH, "//a[contains(text(), 'next') or contains(text(), 'Next')]"),
                (By.XPATH, "//button[contains(text(), '下一页')]"),
                (By.XPATH, "//button[contains(text(), '下页')]"),
                # 通过class查找
                (By.CSS_SELECTOR, "a.next, a.page-next, button.next, button.page-next"),
                (By.XPATH, "//a[@class='next']"),
                (By.XPATH, "//a[@class='page-next']"),
                (By.XPATH, "//a[contains(@class, 'next')]"),
                # 查找分页相关的元素
                (By.XPATH, "//div[contains(@class, 'pagination')]//a[contains(text(), '下一')]"),
                (By.XPATH, "//div[contains(@class, 'page')]//a[contains(text(), '下一')]"),
            ]
            
            for selector_type, selector_value in next_page_selectors:
                try:
                    if selector_type == By.XPATH:
                        elements = self.driver.find_elements(By.XPATH, selector_value)
                    elif selector_type == By.CSS_SELECTOR:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector_value)
                    else:
                        continue
                    
                    for button in elements:
                        if button.is_displayed():
                            # 检查按钮是否被禁用
                            disabled_attr = button.get_attribute("disabled")
                            class_attr = button.get_attribute("class") or ""
                            class_lower = class_attr.lower()
                            
                            # 如果按钮被禁用，跳过
                            if disabled_attr is not None and disabled_attr != "false":
                                continue
                            
                            # 如果class中包含disabled，跳过
                            if "disabled" in class_lower:
                                continue
                            
                            print(f"找到下一页按钮（使用选择器: {selector_type}, {selector_value}）")
                            return button
                except:
                    continue
            
            # 尝试通过JavaScript查找
            print("尝试通过JavaScript查找下一页按钮...")
            button = self.driver.execute_script("""
                var links = document.querySelectorAll('a, button');
                for(var i = 0; i < links.length; i++) {
                    var text = links[i].textContent || links[i].innerText || '';
                    var className = links[i].className || '';
                    if((text.includes('下一页') || text.includes('下页') || 
                        text.includes('next') || text.includes('Next') ||
                        className.includes('next')) && 
                       links[i].offsetParent !== null) {
                        var disabled = links[i].getAttribute('disabled');
                        var classAttr = links[i].className || '';
                        if(disabled === null && !classAttr.toLowerCase().includes('disabled')) {
                            return links[i];
                        }
                    }
                }
                return null;
            """)
            
            if button:
                print("通过JavaScript找到下一页按钮")
                return button
            
            return None
            
        except Exception as e:
            print(f"查找下一页按钮时出错: {e}")
            return None
    
    def click_next_page(self) -> bool:
        """
        点击下一页按钮
        
        Returns:
            是否成功点击并跳转到下一页
        """
        try:
            # 滚动到页面底部，确保分页按钮可见
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            
            button = self.find_next_page_button()
            
            if not button:
                print("未找到下一页按钮，可能已到达最后一页")
                return False
            
            # 记录当前URL
            current_url = self.driver.current_url
            
            # 尝试点击按钮
            try:
                # 滚动到按钮可见
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", button)
                time.sleep(0.5)
                
                # 先尝试普通点击
                button.click()
            except:
                # 如果普通点击失败，尝试JavaScript点击
                self.driver.execute_script("arguments[0].click();", button)
            
            # 等待页面加载
            time.sleep(3)
            
            # 检查是否跳转到新页面
            new_url = self.driver.current_url
            if new_url != current_url:
                print(f"成功跳转到下一页: {new_url}")
                return True
            else:
                # 可能按钮被禁用或者是最后一页
                print("页面未发生变化，可能已到达最后一页")
                return False
                
        except Exception as e:
            print(f"点击下一页按钮时出错: {e}")
            return False
    
    def scroll_page(self, scroll_pause_time: float = 1.0):
        """滚动页面以确保所有内容加载完成（保留原方法用于兼容）"""
        self.scroll_to_bottom(scroll_pause_time)
    
    def save_to_json(self, news_list: List[Dict], filename: str = 'cctv_news.json'):
        """保存结果到JSON文件"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(news_list, f, ensure_ascii=False, indent=2)
        print(f"结果已保存到 {filename}")
    
    def save_to_csv(self, news_list: List[Dict], filename: str = 'cctv_news.csv'):
        """保存结果到CSV文件"""
        if not news_list:
            print("没有数据可保存")
            return
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=['keyword', 'title', 'publish_time', 'summary', 'author', 'url'])
            writer.writeheader()
            writer.writerows(news_list)
        print(f"结果已保存到 {filename}")
    
    def save_to_excel(self, news_list: List[Dict], filename: str = 'cctv_news.xlsx'):
        """保存结果到Excel文件"""
        if not news_list:
            print("没有数据可保存")
            return
        
        try:
            df = pd.DataFrame(news_list, columns=['keyword', 'title', 'publish_time', 'summary', 'author', 'url'])
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='新闻列表')
                
                worksheet = writer.sheets['新闻列表']
                
                # 自动调整列宽
                for idx, col in enumerate(df.columns, start=1):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(col)
                    )
                    adjusted_width = min(max_length + 2, 50)
                    from openpyxl.utils import get_column_letter
                    column_letter = get_column_letter(idx)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 设置标题行样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"结果已保存到 {filename}")
        except Exception as e:
            print(f"保存Excel文件时出错: {e}")
            import traceback
            traceback.print_exc()
    
    def close(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            print("浏览器已关闭")


def main():
    """主函数"""
    import sys
    
    # 定义关键词列表
    keywords = [
        '加班', '996', '劳动法', '加班费', '工作压力', '职场焦虑', '职场压力', 
        '工作焦虑', '工作倦怠', '工作过劳', '工作疲惫', '职业倦怠', '工作躺平',
        '企业责任', '员工关怀', '职业健康', '职场内卷', '职场竞争', '职场PUA', 
        '职场霸凌', '职场心理咨询', '职场心理健康', '职场关怀'
    ]
    
    crawler = None
    all_results = []  # 存储所有关键词的结果
    
    try:
        # 创建爬虫实例
        crawler = CTVTCrawler(headless=False)  # headless=True 隐藏浏览器窗口
        
        print("\n" + "="*80)
        print("开始爬取CCTV新闻")
        print(f"共 {len(keywords)} 个关键词")
        print("将自动爬取所有页面，直到没有下一页")
        print("="*80)
        
        # 遍历每个关键词进行爬取
        for idx, keyword in enumerate(keywords, 1):
            print("\n" + "="*80)
            print(f"【关键词 {idx}/{len(keywords)}】: {keyword}")
            print("="*80)
            
            try:
                # 点击搜索按钮并搜索
                if crawler.click_search_button():
                    if crawler.search_keyword(keyword):
                        # 爬取所有页面（自动下滑、自动翻页，直到没有下一页）
                        news_list = crawler.crawl_all_pages(keyword=keyword)
                        
                        if news_list:
                            print(f"\n关键词 '{keyword}' 共找到 {len(news_list)} 条新闻（爬取结果）")
                            
                            # 筛选2025年的新闻
                            print("正在筛选2025年的新闻...")
                            filtered_news = crawler.filter_by_year(news_list, year=2025)
                            print(f"筛选完成：共 {len(news_list)} 条原始新闻，{len(filtered_news)} 条2025年的新闻")
                            
                            if filtered_news:
                                # 添加到总结果中
                                all_results.extend(filtered_news)
                                print(f"关键词 '{keyword}' 爬取完成，获得 {len(filtered_news)} 条2025年新闻")
                                print(f"当前累计: {len(all_results)} 条新闻")
                            else:
                                print(f"关键词 '{keyword}' 未找到2025年的新闻")
                        else:
                            print(f"关键词 '{keyword}' 未找到任何新闻")
                    else:
                        print(f"关键词 '{keyword}' 搜索失败，跳过")
                else:
                    print(f"关键词 '{keyword}' 无法打开搜索功能，跳过")
                
                # 每个关键词之间稍作停顿
                if idx < len(keywords):
                    print(f"\n等待 2 秒后继续下一个关键词...")
                    time.sleep(2)
                    
            except Exception as e:
                print(f"处理关键词 '{keyword}' 时出错: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # 所有关键词爬取完成后的处理
        print("\n" + "="*80)
        print("所有关键词爬取完成！")
        print("="*80)
        
        if all_results:
            print(f"\n共找到 {len(all_results)} 条2025年的新闻（所有关键词汇总）")
            
            # 打印统计信息
            print("\n" + "="*80)
            print("爬取统计:")
            print("="*80)
            keyword_stats = {}
            for news in all_results:
                kw = news.get('keyword', '未知')
                keyword_stats[kw] = keyword_stats.get(kw, 0) + 1
            
            for kw, count in sorted(keyword_stats.items(), key=lambda x: x[1], reverse=True):
                print(f"  {kw}: {count} 条")
            
            # 打印前10条结果作为预览
            print("\n" + "="*80)
            print("前10条结果预览（2025年）：")
            print("="*80)
            for i, news in enumerate(all_results[:10], 1):
                print(f"\n【新闻 {i}】")
                print(f"关键词: {news.get('keyword', 'N/A')}")
                print(f"标题: {news.get('title', 'N/A')}")
                print(f"发布时间: {news.get('publish_time', 'N/A')}")
                print(f"引言: {news.get('summary', 'N/A')[:100]}..." if len(news.get('summary', '')) > 100 else f"引言: {news.get('summary', 'N/A')}")
                print(f"作者: {news.get('author', 'N/A')}")
                print(f"链接: {news.get('url', 'N/A')}")
            
            if len(all_results) > 10:
                print(f"\n... 还有 {len(all_results) - 10} 条新闻未显示")
            
            # 保存所有结果
            crawler.save_to_excel(all_results, 'cctv_news_all_keywords.xlsx')
            crawler.save_to_csv(all_results, 'cctv_news_all_keywords.csv')
            crawler.save_to_json(all_results, 'cctv_news_all_keywords.json')
            print(f"\n所有结果已保存（包含所有关键词的2025年新闻）")
        else:
            print("未找到任何新闻")
    
    except KeyboardInterrupt:
        print("\n\n用户中断程序")
        # 如果中断，保存已爬取的结果
        if crawler and all_results:
            print("正在保存已爬取的结果...")
            crawler.save_to_excel(all_results, 'cctv_news_interrupted.xlsx')
            crawler.save_to_csv(all_results, 'cctv_news_interrupted.csv')
            crawler.save_to_json(all_results, 'cctv_news_interrupted.json')
            print(f"已保存 {len(all_results)} 条结果到中断文件")
    except Exception as e:
        print(f"程序运行出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if crawler:
            crawler.close()


if __name__ == "__main__":
    main()
